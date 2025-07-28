import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import io
from typing import Optional, List


class ExcelConsolidator:
    def __init__(self):
        self.input_file = None
        self.output_file = None

    def find_header_row(self, df_raw):
        """Find the row that contains 'Fecha' and 'Asiento' headers"""
        for idx, row in df_raw.iterrows():
            row_str = [str(val).lower() if pd.notna(val) else "" for val in row]
            has_fecha = any("fecha" in val for val in row_str)
            has_asiento = any("asiento" in val for val in row_str)

            if has_fecha and has_asiento:
                return idx
        return None

    def process_excel_file(self, input_file_bytes, input_filename, progress_callback=None):
        """Process the uploaded Excel file and return the consolidated workbook"""

        # Create Excel file from bytes
        excel_file = pd.ExcelFile(io.BytesIO(input_file_bytes))

        # Create a new workbook for output
        output_workbook = openpyxl.Workbook()
        output_workbook.remove(output_workbook.active)  # Remove default sheet

        master_data = []

        # Process each sheet
        for sheet_name in excel_file.sheet_names:
            if progress_callback:
                progress_callback(f"Procesando hoja: **{sheet_name}**")

            # Read the sheet without headers first to find the data structure
            df_raw = pd.read_excel(
                io.BytesIO(input_file_bytes), sheet_name=sheet_name, header=None
            )

            # Copy original sheet to output
            ws = output_workbook.create_sheet(title=sheet_name)
            for r in dataframe_to_rows(df_raw, index=False, header=False):
                ws.append(r)

            # Find the header row (look for "Fecha" and "Asiento")
            header_row = self.find_header_row(df_raw)

            if header_row is not None:
                # Read the sheet again with the correct header row
                df = pd.read_excel(
                    io.BytesIO(input_file_bytes),
                    sheet_name=sheet_name,
                    header=header_row,
                )

                # Filter rows that have values in both Fecha and Asiento columns
                fecha_col = None
                asiento_col = None

                for col in df.columns:
                    if isinstance(col, str) and "fecha" in col.lower():
                        fecha_col = col
                    if isinstance(col, str) and "asiento" in col.lower():
                        asiento_col = col

                if fecha_col and asiento_col:
                    # Forward fill fecha and asiento for continuation rows
                    df[fecha_col] = df[fecha_col].ffill()
                    df[asiento_col] = df[asiento_col].ffill()

                    # Remove completely empty rows
                    valid_rows = df.dropna(how="all")

                    # Filter rows that have meaningful data beyond just fecha/asiento
                    # A valid entry should have at least one non-null value in other important columns
                    important_cols = [col for col in valid_rows.columns 
                                    if col not in [fecha_col, asiento_col] and 
                                    col not in ['Sheet_Origin']]
                    
                    # Keep rows that have at least one non-null value in important columns
                    meaningful_data_mask = valid_rows[important_cols].notna().any(axis=1)
                    valid_rows = valid_rows[meaningful_data_mask]

                    # Filter out total/summary rows (usually at the end)
                    # Check if any column contains words indicating it's a total row
                    total_keywords = ["total", "suma", "suman", "totales", "resumen"]
                    mask = True
                    for col in valid_rows.columns:
                        if valid_rows[col].dtype == "object":  # String columns
                            col_mask = ~valid_rows[col].astype(
                                str
                            ).str.lower().str.contains(
                                "|".join(total_keywords), na=False
                            )
                            mask = mask & col_mask

                    valid_rows = valid_rows[mask]

                    if not valid_rows.empty:
                        # Add sheet name as a column for identification
                        valid_rows = valid_rows.copy()
                        valid_rows["Sheet_Origin"] = sheet_name
                        master_data.append(valid_rows)
                        if progress_callback:
                            progress_callback(
                                f"✅ Se encontraron {len(valid_rows)} entradas válidas en {sheet_name}"
                            )
                    else:
                        if progress_callback:
                            progress_callback(
                                f"⚠️ No se encontraron entradas válidas en {sheet_name}"
                            )
                else:
                    if progress_callback:
                        progress_callback(f"⚠️ Columnas requeridas no encontradas en {sheet_name}")
            else:
                if progress_callback:
                    progress_callback(f"⚠️ Fila de encabezados no encontrada en {sheet_name}")

        # Create master sheet
        if master_data:
            if progress_callback:
                progress_callback("Creando hoja maestra...")

            master_df = pd.concat(master_data, ignore_index=True)

            # Create master sheet
            master_ws = output_workbook.create_sheet(title="MASTER")
            for r in dataframe_to_rows(master_df, index=False, header=True):
                master_ws.append(r)

            # Move master sheet to the beginning
            output_workbook.move_sheet(
                "MASTER", offset=-len(output_workbook.sheetnames) + 1
            )

            if progress_callback:
                progress_callback(f"✅ Hoja maestra creada con {len(master_df)} entradas totales")

            return output_workbook, master_df
        else:
            if progress_callback:
                progress_callback("¡No se encontraron datos válidos en ninguna hoja!")
            return None, None