import pytest
import pandas as pd
import openpyxl
from pathlib import Path
import os
import sys
import tempfile

# Add parent directory to path to import consolidator
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

from app.consolidator import ExcelConsolidator


class TestExcelConsolidator:
    def setup_method(self):
        self.consolidator = ExcelConsolidator()
        self.test_file_path = Path(__file__).parent.parent / "data" / "EMPRESA_1-A3_TODO_EJERCICIO_2020.xlsx"
        self.temp_files = []

    def teardown_method(self):
        """Clean up any temporary files created during tests"""
        for temp_file in self.temp_files:
            if temp_file.exists():
                temp_file.unlink()

    def test_file_exists(self):
        """Test that the required test file exists"""
        assert self.test_file_path.exists(), f"Test file {self.test_file_path} should exist in tests/data directory"

    def test_process_excel_file_with_real_file(self):
        """Test processing the actual Excel file from data directory"""
        if not self.test_file_path.exists():
            pytest.skip(f"Test file {self.test_file_path} not found")

        # Read the test file
        with open(self.test_file_path, "rb") as f:
            file_bytes = f.read()

        # Capture progress messages
        progress_messages = []

        def capture_progress(message):
            progress_messages.append(message)

        # Process the file
        output_workbook, master_df = self.consolidator.process_excel_file(
            file_bytes, self.test_file_path.name, progress_callback=capture_progress
        )

        # Assertions
        assert output_workbook is not None, "Output workbook should not be None"
        assert master_df is not None, "Master dataframe should not be None"
        assert len(master_df) > 0, "Master dataframe should contain data"

        # Check that MASTER sheet exists
        sheet_names = output_workbook.sheetnames
        assert "MASTER" in sheet_names, "MASTER sheet should exist in output workbook"
        assert sheet_names[0] == "MASTER", "MASTER sheet should be the first sheet"

        # Check that master dataframe has expected columns
        assert (
            "Sheet_Origin" in master_df.columns
        ), "Master dataframe should have Sheet_Origin column"

        # Check for Fecha and Asiento columns (at least one should exist)
        fecha_cols = [
            col
            for col in master_df.columns
            if isinstance(col, str) and "fecha" in col.lower()
        ]
        asiento_cols = [
            col
            for col in master_df.columns
            if isinstance(col, str) and "asiento" in col.lower()
        ]

        assert (
            len(fecha_cols) > 0
        ), "Master dataframe should have at least one Fecha column"
        assert (
            len(asiento_cols) > 0
        ), "Master dataframe should have at least one Asiento column"

        # Check that progress messages were captured
        assert len(progress_messages) > 0, "Progress messages should be captured"

        # Verify the corrected counting logic - should have reasonable number of entries
        assert len(master_df) < 50, "Should have reasonable number of entries (not inflated by empty rows)"
        
        # Print some info for debugging
        print(f"Processed {len(master_df)} total entries")
        print(f"Sheets processed: {master_df['Sheet_Origin'].unique()}")
        print(f"Columns in master dataframe: {list(master_df.columns)}")

    def test_process_excel_file_creates_master_sheet(self):
        """Test that the MASTER sheet is created with correct structure"""
        if not self.test_file_path.exists():
            pytest.skip(f"Test file {self.test_file_path} not found")

        with open(self.test_file_path, "rb") as f:
            file_bytes = f.read()

        output_workbook, master_df = self.consolidator.process_excel_file(
            file_bytes, self.test_file_path.name
        )

        # Get the MASTER sheet
        master_sheet = output_workbook["MASTER"]

        # Check that the sheet has data (header + at least one data row)
        assert master_sheet.max_row > 1, "MASTER sheet should have header + data rows"
        assert master_sheet.max_column > 0, "MASTER sheet should have columns"

        # Check that the first row contains headers
        first_row = [cell.value for cell in master_sheet[1]]
        assert any(
            cell for cell in first_row if cell is not None
        ), "First row should contain headers"

    def test_process_excel_file_saves_correctly(self):
        """Test that the processed file can be saved and reloaded correctly"""
        if not self.test_file_path.exists():
            pytest.skip(f"Test file {self.test_file_path} not found")

        with open(self.test_file_path, "rb") as f:
            file_bytes = f.read()

        output_workbook, master_df = self.consolidator.process_excel_file(
            file_bytes, self.test_file_path.name
        )

        # Save to a temporary file
        temp_output = Path(tempfile.mktemp(suffix='.xlsx'))
        self.temp_files.append(temp_output)
        
        output_workbook.save(temp_output)
        assert temp_output.exists(), "Output file should be created"

        # Verify we can reload the saved file
        reloaded_df = pd.read_excel(temp_output, sheet_name="MASTER")
        assert len(reloaded_df) == len(master_df), "Reloaded data should match original"
        assert list(reloaded_df.columns) == list(master_df.columns), "Columns should match"

    def test_process_excel_file_correct_entry_count(self):
        """Test that the entry counting logic correctly excludes empty forward-filled rows"""
        if not self.test_file_path.exists():
            pytest.skip(f"Test file {self.test_file_path} not found")

        with open(self.test_file_path, "rb") as f:
            file_bytes = f.read()

        progress_messages = []
        def capture_progress(message):
            progress_messages.append(message)

        output_workbook, master_df = self.consolidator.process_excel_file(
            file_bytes, self.test_file_path.name, progress_callback=capture_progress
        )

        # Check that January doesn't have 174 entries (the bug we fixed)
        enero_messages = [msg for msg in progress_messages if "Enero" in msg and "entradas válidas" in msg]
        assert len(enero_messages) == 1, "Should have exactly one message about January entries"
        
        enero_message = enero_messages[0]
        # Extract the number from the message
        import re
        match = re.search(r'(\d+) entradas válidas en Enero', enero_message)
        assert match is not None, "Should be able to extract entry count from message"
        
        entry_count = int(match.group(1))
        assert entry_count < 20, f"January should have reasonable number of entries, got {entry_count}"
        assert entry_count > 0, "January should have at least some entries"